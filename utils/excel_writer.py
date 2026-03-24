"""utils/excel_writer.py — Excel 問卷自動填寫模組

負責將辨識結果自動填寫至「成大第一階段辨識問卷」Excel 檔案。
支援 .xlsx 和 .xlsm 格式。

Excel 欄位對應（第 3 列為標題列，資料從第 4 列開始）：
    A: 藥盤序號
    B: 階段別（固定「第1階段」）
    C: 測試日期
    D: 測試人員/紀錄者（留空）
    E: 情境（留空）
    F: 條件(光源角度)（留空）
    G: 藥盤整體判定結果(自動判斷)
    H: 品項種類
    I: 總量
    J: 品項數量
    K: 問題描述（留空或自動填入）
    L: 備註（留空）
    M: raw data（圖片路徑）
    N~: Pill_1, Pill_2, ... (動態欄位，有幾顆填幾個)
"""

from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelWriter:
    """Excel 問卷填寫器
    
    使用方式：
        writer = ExcelWriter("問卷.xlsx")
        writer.write_verification_data(
            tray_id="000001",
            timestamp="2026-03-24 10:30:00",
            variety_count=3,
            variety_correct=True,
            total_count=10,
            total_correct=True,
            pills=[...],
            name_answers=[True, False, ...],
            dose_answers=[True, True, ...]
        )
        writer.save()
    """
    
    def __init__(self, excel_path: str | Path, sheet_name: Optional[str] = None):
        """初始化 Excel 寫入器
        
        Args:
            excel_path: Excel 檔案路徑
            sheet_name: 工作表名稱（預設使用 active sheet）
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl 未安裝，請執行: pip install openpyxl")
        
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel 檔案不存在: {self.excel_path}")
        
        # 載入工作簿（保留巨集）
        self.workbook = load_workbook(self.excel_path, keep_vba=True)
        
        # 選擇工作表
        if sheet_name:
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {self.workbook.sheetnames}")
            self.sheet = self.workbook[sheet_name]
        else:
            self.sheet = self.workbook.active
        
        print(f"[excel] 已載入: {self.excel_path.name}, 工作表: {self.sheet.title}")
    
    def find_next_empty_row(self, start_row: int = 4, column: int = 1) -> int:
        """找到下一個空白列（用於追加資料）
        
        Args:
            start_row: 開始搜尋的列號（預設從第 4 列開始，第 3 列是標題）
            column: 檢查的欄位（預設第 1 欄 - 藥盤序號）
        
        Returns:
            空白列的列號
        """
        row = start_row
        while self.sheet.cell(row, column).value is not None:
            row += 1
        return row
    
    def write_verification_data(
        self,
        tray_id: str,
        timestamp: str,
        variety_count: int,
        variety_correct: bool | None,
        total_count: int,
        total_correct: bool | None,
        pills: list,
        name_answers: list[bool | None],
        dose_answers: list[list[bool | None]],  # 每個類別下的每顆藥錠
        image_path: str = "",
        start_row: Optional[int] = None,
    ):
        """將驗證資料寫入 Excel（按照問卷實際欄位結構）
        
        Args:
            tray_id: 藥盤序號
            timestamp: 時間戳記（格式：YYYY-MM-DD HH:MM:SS）
            variety_count: 品項種類數量（不重複藥品數）
            variety_correct: 品項種類驗證結果
            total_count: 總顆數
            total_correct: 總量驗證結果
            pills: PillEntry 列表（每顆藥一筆）
            name_answers: 各藥品類別的名稱驗證結果（per category）
            dose_answers: 各藥品類別下每顆藥錠的劑量驗證結果（per category, per pill）
            image_path: 圖片路徑（相對於問卷檔案）
            start_row: 寫入的起始列（None 則自動找下一個空白列）
        """
        if start_row is None:
            row = self.find_next_empty_row()
        else:
            row = start_row
        
        # A 欄：藥盤序號
        self.sheet.cell(row, 1, tray_id)
        
        # B 欄：階段別（固定值）
        self.sheet.cell(row, 2, "第1階段")
        
        # C 欄：測試日期（從 timestamp 取日期部分）
        try:
            dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
            date_str = dt.strftime("%Y/%m/%d")
        except:
            date_str = timestamp.split()[0] if " " in timestamp else timestamp
        self.sheet.cell(row, 3, date_str)
        
        # D~F 欄：測試人員、情境、條件 — 留空（由人工填寫）
        # self.sheet.cell(row, 4, "")  # 測試人員/紀錄者
        # self.sheet.cell(row, 5, "")  # 情境
        # self.sheet.cell(row, 6, "")  # 條件(光源角度)
        
        # 展開所有 dose_answers 為一維列表（用於整體判斷）
        all_dose_answers = [ans for cat_answers in dose_answers for ans in cat_answers]
        
        # G 欄：藥盤整體判定結果(自動判斷)
        # 只有所有驗證項目都是 True 才填「正確」
        all_correct = (
            variety_correct is True and
            total_correct is True and
            all(ans is True for ans in name_answers) and
            all(ans is True for ans in all_dose_answers)
        )
        self.sheet.cell(row, 7, "正確" if all_correct else "錯誤")
        
        # H 欄：品項種類（正確/錯誤）
        self.sheet.cell(row, 8, self._bool_to_text(variety_correct))
        
        # I 欄：總量（正確/錯誤）
        self.sheet.cell(row, 9, self._bool_to_text(total_correct))
        
        # J 欄：品項數量（正確/錯誤）
        # 根據範例，這也是填「正確/錯誤」，暫時與 variety_correct 相同
        # TODO: 確認 J 欄的實際意義
        self.sheet.cell(row, 10, self._bool_to_text(variety_correct))
        
        # K 欄：問題描述 — 留空（不自動填寫）
        # L 欄：備註 — 留空
        # self.sheet.cell(row, 12, "")
        
        # M 欄：raw data — 圖片路徑
        if image_path:
            self.sheet.cell(row, 13, image_path)
        
        # N 欄起：Pill_1, Pill_2, ... — 有幾顆填幾個欄位
        for i, pill in enumerate(pills):
            col = 14 + i  # N 欄是第 14 欄
            self.sheet.cell(row, col, pill.name or "未識別")
        
        print(f"[excel] 資料已寫入第 {row} 列（共 {len(pills)} 顆藥品）")
        return row
    
    def _generate_problem_description(
        self,
        variety_correct: bool | None,
        total_correct: bool | None,
        name_answers: list[bool | None],
        dose_answers: list[list[bool | None]],
        pills: list,
    ) -> str:
        """自動產生問題描述（僅在有錯誤時）"""
        problems = []
        
        if variety_correct is False:
            problems.append("品項種類錯誤")
        if total_correct is False:
            problems.append("總量錯誤")
        
        # 根據新的資料結構處理：name_answers 是 per category, dose_answers 是 per category per pill
        # pills 列表保持 flat（所有藥錠）
        pill_idx = 0
        for cat_idx, (name_ans, cat_dose_answers) in enumerate(zip(name_answers, dose_answers)):
            if name_ans is False:
                # 取該類別的第一顆藥錠名稱
                if pill_idx < len(pills):
                    pill_name = pills[pill_idx].name or "未識別"
                    problems.append(f"類別{cat_idx+1}({pill_name})名稱錯誤")
            
            for pill_in_cat_idx, dose_ans in enumerate(cat_dose_answers):
                if dose_ans is False:
                    actual_pill_idx = pill_idx + pill_in_cat_idx
                    if actual_pill_idx < len(pills):
                        pill = pills[actual_pill_idx]
                        problems.append(f"{pill.full_label or f'第{actual_pill_idx+1}顆'}({pill.name or '未識別'})劑量錯誤")
            
            pill_idx += len(cat_dose_answers)
        
        return "；".join(problems) if problems else ""
    
    @staticmethod
    def _bool_to_text(value: bool | None) -> str:
        """將布林值轉換為文字"""
        if value is True:
            return "正確"
        elif value is False:
            return "錯誤"
        else:
            return "未填"
    
    def save(self, output_path: Optional[str | Path] = None):
        """儲存 Excel 檔案
        
        Args:
            output_path: 輸出路徑（None 則覆寫原檔案）
        """
        save_path = Path(output_path) if output_path else self.excel_path
        self.workbook.save(save_path)
        print(f"[excel] 已儲存: {save_path}")
    
    def close(self):
        """關閉工作簿"""
        self.workbook.close()


def create_backup(excel_path: str | Path) -> Path:
    """建立 Excel 檔案備份
    
    Args:
        excel_path: 原始檔案路徑
    
    Returns:
        備份檔案路徑
    """
    excel_path = Path(excel_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = excel_path.parent / f"{excel_path.stem}_backup_{timestamp}{excel_path.suffix}"
    
    import shutil
    shutil.copy2(excel_path, backup_path)
    print(f"[excel] 已建立備份: {backup_path.name}")
    return backup_path


__all__ = ["ExcelWriter", "create_backup", "HAS_OPENPYXL"]
